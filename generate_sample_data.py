import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import random, datetime

wb = Workbook()
random.seed(42)

hf = Font(bold=True, color="FFFFFF", size=11)
hfill = PatternFill("solid", start_color="1877F2")
ctr = Alignment(horizontal="center", vertical="center")

def hrow(ws, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = hf; c.fill = hfill; c.alignment = ctr

def colw(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

OFFICES = [
    "National Sales Team","Bay Area","Arizona","Sacramento","Pasadena",
    "Orange County","CDR","Las Vegas","Corporate","Fresno",
    "San Antonio","San Diego","Ventura/SB County"
]
CAMPAIGNS = ["OUTCOME_SALES","CONVERSIONS","OUTCOME_LEADS","OUTCOME_TRAFFIC","OUTCOME_ENGAGEMENT"]
CAMP_COLORS = {"OUTCOME_SALES":"1877F2","CONVERSIONS":"10b981","OUTCOME_LEADS":"f59e0b","OUTCOME_TRAFFIC":"8b5cf6","OUTCOME_ENGAGEMENT":"06b6d4"}

# ── Sheet 1: Overview (MTD) ────────────────────────────────────────────────────
ws1 = wb.active; ws1.title = "Overview"
hrow(ws1, ["Metric","Current Period","Previous Period","Change %"])
overview_data = [
    ["Spend ($)",        112300,  128700, -12.7],
    ["Clicks",           263900,  305700, -13.7],
    ["Conversions",        2700,    3546, -23.8],
    ["CRM Leads",          3100,    4600, -32.4],
    ["Appointments",        480,     390,  23.1],
    ["Customers",           210,     175,  20.0],
    ["Sales Amount ($)", 560000,  430000,  30.2],
]
for r in overview_data: ws1.append(r)
colw(ws1, [22,18,18,12])

# ── Sheet 2: Campaign Performance ─────────────────────────────────────────────
ws2 = wb.create_sheet("Campaign Performance")
hdrs2 = ["Campaign Objective","Spend ($)","Impressions","CPM ($)","Clicks","CPC ($)",
         "CRM Leads","Conversions","Appointments","Customers","Sales Amount ($)","ROAS","CTR (%)"]
hrow(ws2, hdrs2)
camp_data = [
    ["OUTCOME_SALES",      90900, 16100000, 5.64, 201300, 0.45, 665, 2000, 210, 95, 448000, 25.6, 1.25],
    ["CONVERSIONS",        17200,  2800000, 6.14,  35300, 0.49, 316,  717, 180, 80,  95000,  0.8, 1.26],
    ["OUTCOME_LEADS",        931,   271600, 3.43,   6000, 0.16,2200,    0,  60, 25,  17000, -1.0, 2.21],
    ["OUTCOME_TRAFFIC",     1200,   331700, 3.62,  14100, 0.09,   0,    0,  20,  5,      0, -1.0, 4.25],
    ["OUTCOME_ENGAGEMENT",  1900,   982600, 1.93,   7200, 0.26,   2,    0,  10,  5,      0, -1.0, 0.73],
]
for r in camp_data: ws2.append(r)
colw(ws2, [22,12,14,10,10,10,10,12,13,12,16,10,10])

# ── Sheet 3: Daily Performance ────────────────────────────────────────────────
ws3 = wb.create_sheet("Daily Performance")
hdrs3 = ["Date","Campaign","Spend ($)","Impressions","Reach","Clicks","CRM Leads",
         "Conversions","Appointments","Customers","Sales Amount ($)","ROAS","Frequency"]
hrow(ws3, hdrs3)
base = datetime.date(2024, 12, 1)
for i in range(31):
    d = base + datetime.timedelta(days=i)
    for camp in CAMPAIGNS:
        mult = {"OUTCOME_SALES":1.0,"CONVERSIONS":0.19,"OUTCOME_LEADS":0.01,"OUTCOME_TRAFFIC":0.013,"OUTCOME_ENGAGEMENT":0.021}[camp]
        ws3.append([
            d.strftime("%Y-%m-%d"), camp,
            round(random.uniform(2800,4500)*mult, 2),
            round(random.uniform(500000,900000)*mult),
            round(random.uniform(140000,220000)*mult),
            round(random.uniform(6000,12000)*mult),
            round(random.uniform(60,150)*mult) if camp in ["OUTCOME_SALES","CONVERSIONS","OUTCOME_LEADS"] else 0,
            round(random.uniform(50,120)*mult) if camp in ["OUTCOME_SALES","CONVERSIONS"] else 0,
            round(random.uniform(5,15)*mult),
            round(random.uniform(2,8)*mult),
            round(random.uniform(12000,28000)*mult, 2),
            round(random.uniform(3.5,8.0), 2) if camp in ["OUTCOME_SALES","CONVERSIONS"] else 0,
            round(random.uniform(12,18), 1),
        ])
colw(ws3, [14,22,12,14,12,10,12,13,13,12,18,8,12])

# ── Sheet 4: Territory Performance ───────────────────────────────────────────
ws4 = wb.create_sheet("Territory Performance")
hdrs4 = ["Territory","Campaign","Unique Leads","New Leads","Appointments","Quote",
         "Customers","Sales Amount ($)","NL Customers","NL Sales ($)",
         "Leads %","Sales %","APT/Leads","Order/APT","Order/Leads","Spend ($)","ROAS"]
hrow(ws4, hdrs4)

office_weights = {
    "National Sales Team":0.34,"Bay Area":0.18,"Arizona":0.11,"Sacramento":0.10,
    "Pasadena":0.07,"Orange County":0.05,"CDR":0.03,"Las Vegas":0.03,
    "Corporate":0.02,"Fresno":0.02,"San Antonio":0.02,"San Diego":0.02,"Ventura/SB County":0.01,
}
total_leads = 3100; total_sales = 560000; total_spend = 112300

for office, wt in office_weights.items():
    for camp in CAMPAIGNS:
        cwt = {"OUTCOME_SALES":0.55,"CONVERSIONS":0.25,"OUTCOME_LEADS":0.12,"OUTCOME_TRAFFIC":0.05,"OUTCOME_ENGAGEMENT":0.03}[camp]
        ul  = max(0, round(total_leads * wt * cwt * random.uniform(0.85,1.15)))
        nl  = round(ul * random.uniform(0.85, 1.0))
        apt = round(ul * random.uniform(0.1, 0.5))
        quote = round(apt * random.uniform(0.3, 0.8))
        cust  = round(quote * random.uniform(0.2, 0.6)) if camp in ["OUTCOME_SALES","CONVERSIONS"] else 0
        sales = round(cust * random.uniform(4000, 8000), 2)
        nl_cust = round(cust * random.uniform(0.4, 0.8))
        nl_sales = round(nl_cust * random.uniform(4000, 8000), 2)
        leads_pct = round(ul / total_leads * 100, 2) if total_leads else 0
        sales_pct = round(sales / total_sales * 100, 2) if total_sales else 0
        apt_leads = round(apt / ul * 100, 2) if ul else 0
        ord_apt   = round(cust / apt * 100, 2) if apt else 0
        ord_leads = round(cust / ul * 100, 2) if ul else 0
        spend = round(total_spend * wt * cwt * random.uniform(0.85, 1.15), 2)
        roas  = round(sales / spend, 2) if spend and sales else 0
        ws4.append([office, camp, ul, nl, apt, quote, cust, sales, nl_cust, nl_sales,
                    leads_pct, sales_pct, apt_leads, ord_apt, ord_leads, spend, roas])

colw(ws4, [20,22,13,11,13,8,11,16,14,13,10,10,12,12,13,12,8])

wb.save("meta_ads_data.xlsx")
print("Done — meta_ads_data.xlsx created with 4 sheets.")
